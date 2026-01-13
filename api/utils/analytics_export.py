"""
数据分析报告导出工具

支持导出格式：
- PDF (.pdf)
- Excel (.xlsx)

支持合并多个报告到一个文件
"""
import io
from datetime import datetime, date
from typing import Optional, Dict, Any, List

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class AnalyticsExporter:
    """数据分析报告导出器"""

    def __init__(self):
        self.styles = getSampleStyleSheet()
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=20,
            alignment=TA_CENTER,
        )
        self.section_title_style = ParagraphStyle(
            'SectionTitle',
            parent=self.styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=15,
            spaceBefore=10,
        )

    def _create_table_style(self, header_color: str = '#3B82F6'):
        """创建统一的表格样式"""
        return TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(header_color)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F9FAFB')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB')),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ])

    def _add_overview_section(self, story: List, overview_data: Dict[str, Any]):
        """添加概览部分到PDF"""
        story.append(Paragraph("Overview", self.section_title_style))
        story.append(Spacer(1, 0.1 * inch))

        table_data = [
            ["Metric", "Value"],
            ["Total Cases", str(overview_data.get('total_cases', 0))],
            ["Total Diagnoses", str(overview_data.get('total_diagnoses', 0))],
            ["Avg Execution Time", f"{overview_data.get('avg_execution_time_ms', 0) / 1000:.2f}s"],
            ["Completion Rate", f"{overview_data.get('completion_rate', 0):.1f}%"],
            ["Active Users", str(overview_data.get('active_users', 0))],
        ]

        table = Table(table_data, colWidths=[3 * inch, 2 * inch])
        table.setStyle(self._create_table_style('#3B82F6'))
        story.append(table)
        story.append(Spacer(1, 0.3 * inch))

    def _add_demographics_section(self, story: List, demographics_data: Dict[str, Any]):
        """添加人口统计部分到PDF"""
        story.append(Paragraph("Demographics", self.section_title_style))
        story.append(Spacer(1, 0.1 * inch))

        # 年龄分布
        story.append(Paragraph("Age Distribution", self.styles['Heading3']))
        story.append(Spacer(1, 0.05 * inch))

        age_data = demographics_data.get('age_distribution', [])
        if age_data:
            table_data = [["Age Group", "Count"]]
            for item in age_data:
                table_data.append([item.get('age_group', ''), str(item.get('count', 0))])

            table = Table(table_data, colWidths=[2 * inch, 1.5 * inch])
            table.setStyle(self._create_table_style('#3B82F6'))
            story.append(table)

        story.append(Spacer(1, 0.2 * inch))

        # 性别分布
        story.append(Paragraph("Gender Distribution", self.styles['Heading3']))
        story.append(Spacer(1, 0.05 * inch))

        gender_data = demographics_data.get('gender_distribution', [])
        if gender_data:
            table_data = [["Gender", "Count"]]
            gender_map = {'male': 'Male', 'female': 'Female', 'other': 'Other', 'unknown': 'Unknown'}
            for item in gender_data:
                gender = item.get('gender', '')
                display_gender = gender_map.get(gender.lower(), gender)
                table_data.append([display_gender, str(item.get('count', 0))])

            table = Table(table_data, colWidths=[2 * inch, 1.5 * inch])
            table.setStyle(self._create_table_style('#10B981'))
            story.append(table)

        story.append(Spacer(1, 0.3 * inch))

    def _add_trends_section(self, story: List, case_trends: Dict[str, Any],
                           diagnosis_trends: Dict[str, Any], performance_data: Dict[str, Any]):
        """添加趋势部分到PDF"""
        story.append(Paragraph("Trends & Performance", self.section_title_style))
        story.append(Spacer(1, 0.1 * inch))

        # 病例趋势
        story.append(Paragraph("Case Trends (Recent 15)", self.styles['Heading3']))
        story.append(Spacer(1, 0.05 * inch))

        case_series = case_trends.get('series', [])
        if case_series:
            table_data = [["Date", "Cases"]]
            for item in case_series[-15:]:
                table_data.append([item.get('date', ''), str(item.get('count', 0))])

            table = Table(table_data, colWidths=[2 * inch, 1.5 * inch])
            table.setStyle(self._create_table_style('#3B82F6'))
            story.append(table)

        story.append(Spacer(1, 0.2 * inch))

        # 性能统计
        story.append(Paragraph("Performance Statistics", self.styles['Heading3']))
        story.append(Spacer(1, 0.05 * inch))

        exec_stats = performance_data.get('execution_time_stats', {})
        if exec_stats:
            table_data = [
                ["Metric", "Value"],
                ["Min Time", f"{exec_stats.get('min', 0) / 1000:.2f}s"],
                ["Max Time", f"{exec_stats.get('max', 0) / 1000:.2f}s"],
                ["Avg Time", f"{exec_stats.get('avg', 0) / 1000:.2f}s"],
                ["Median Time", f"{exec_stats.get('median', 0) / 1000:.2f}s"],
                ["Total Diagnoses", str(performance_data.get('total_diagnoses', 0))],
            ]

            table = Table(table_data, colWidths=[2.5 * inch, 1.5 * inch])
            table.setStyle(self._create_table_style('#10B981'))
            story.append(table)

        story.append(Spacer(1, 0.3 * inch))

    def _add_models_section(self, story: List, model_data: Dict[str, Any]):
        """添加模型对比部分到PDF"""
        story.append(Paragraph("Model Comparison", self.section_title_style))
        story.append(Spacer(1, 0.1 * inch))

        models = model_data.get('models', [])
        if models:
            total_usage = sum(m.get('usage_count', 0) for m in models)
            table_data = [["Model Name", "Usage", "Avg Time", "Usage %"]]
            for model in models:
                usage = model.get('usage_count', 0)
                percentage = (usage / total_usage * 100) if total_usage > 0 else 0
                table_data.append([
                    model.get('model_name', ''),
                    str(usage),
                    f"{model.get('avg_execution_time_ms', 0) / 1000:.2f}s",
                    f"{percentage:.1f}%"
                ])

            table = Table(table_data, colWidths=[2.5 * inch, 1 * inch, 1 * inch, 1 * inch])
            table.setStyle(self._create_table_style('#8B5CF6'))
            story.append(table)

        story.append(Spacer(1, 0.3 * inch))

    def _add_users_section(self, story: List, user_data: Dict[str, Any]):
        """添加用户活动部分到PDF"""
        story.append(Paragraph("User Activity", self.section_title_style))
        story.append(Spacer(1, 0.1 * inch))

        top_creators = user_data.get('top_creators', [])
        if top_creators:
            table_data = [["Rank", "User", "Cases", "Diagnoses", "Score"]]
            for i, user in enumerate(top_creators[:15], 1):
                score = user.get('case_count', 0) + user.get('diagnosis_count', 0) * 2
                table_data.append([
                    str(i),
                    user.get('full_name') or user.get('username', ''),
                    str(user.get('case_count', 0)),
                    str(user.get('diagnosis_count', 0)),
                    str(score)
                ])

            table = Table(table_data, colWidths=[0.5 * inch, 2 * inch, 1 * inch, 1 * inch, 1 * inch])
            table.setStyle(self._create_table_style('#F59E0B'))
            story.append(table)

        story.append(Spacer(1, 0.3 * inch))

    def export_combined_pdf(
        self,
        report_types: List[str],
        data: Dict[str, Any],
        start_date: Optional[date] = None,
        end_date: Optional[date] = None
    ) -> bytes:
        """
        导出合并的PDF报告

        Args:
            report_types: 要导出的报告类型列表 ['overview', 'demographics', 'trends', 'models', 'users']
            data: 包含各类型数据的字典
            start_date: 开始日期
            end_date: 结束日期
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=50,
            leftMargin=50,
            topMargin=50,
            bottomMargin=30,
        )

        story = []

        # 添加报告标题
        story.append(Paragraph("Data Analytics Report", self.title_style))
        story.append(Spacer(1, 0.1 * inch))

        # 添加日期范围和生成时间
        date_info = []
        if start_date or end_date:
            date_range = f"Date Range: {start_date or 'N/A'} to {end_date or 'N/A'}"
            date_info.append(date_range)
        date_info.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        date_info.append(f"Sections: {', '.join([t.title() for t in report_types])}")

        for info in date_info:
            story.append(Paragraph(info, self.styles['Normal']))

        story.append(Spacer(1, 0.3 * inch))
        story.append(Paragraph("_" * 70, self.styles['Normal']))
        story.append(Spacer(1, 0.2 * inch))

        # 根据选择的报告类型添加各部分
        for i, report_type in enumerate(report_types):
            if i > 0:
                story.append(PageBreak())

            if report_type == 'overview' and 'overview' in data:
                self._add_overview_section(story, data['overview'])
            elif report_type == 'demographics' and 'demographics' in data:
                self._add_demographics_section(story, data['demographics'])
            elif report_type == 'trends' and 'trends' in data:
                trends_data = data['trends']
                self._add_trends_section(
                    story,
                    trends_data.get('case_trends', {}),
                    trends_data.get('diagnosis_trends', {}),
                    trends_data.get('performance', {})
                )
            elif report_type == 'models' and 'models' in data:
                self._add_models_section(story, data['models'])
            elif report_type == 'users' and 'users' in data:
                self._add_users_section(story, data['users'])

        doc.build(story)
        pdf_bytes = buffer.getvalue()
        buffer.close()
        return pdf_bytes

    def export_combined_excel(
        self,
        report_types: List[str],
        data: Dict[str, Any],
        start_date: Optional[date] = None,
        end_date: Optional[date] = None
    ) -> bytes:
        """
        导出合并的Excel报告

        Args:
            report_types: 要导出的报告类型列表
            data: 包含各类型数据的字典
            start_date: 开始日期
            end_date: 结束日期
        """
        wb = Workbook()

        # 样式定义
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        first_sheet = True

        # 根据选择的报告类型添加各Sheet
        for report_type in report_types:
            if report_type == 'overview' and 'overview' in data:
                if first_sheet:
                    ws = wb.active
                    ws.title = "Overview"
                    first_sheet = False
                else:
                    ws = wb.create_sheet("Overview")
                self._add_overview_sheet(ws, data['overview'], header_font, border, start_date, end_date)

            elif report_type == 'demographics' and 'demographics' in data:
                if first_sheet:
                    ws = wb.active
                    ws.title = "Demographics"
                    first_sheet = False
                else:
                    ws = wb.create_sheet("Demographics")
                self._add_demographics_sheet(ws, data['demographics'], header_font, border)

            elif report_type == 'trends' and 'trends' in data:
                if first_sheet:
                    ws = wb.active
                    ws.title = "Trends"
                    first_sheet = False
                else:
                    ws = wb.create_sheet("Trends")
                trends_data = data['trends']
                self._add_trends_sheet(
                    ws,
                    trends_data.get('case_trends', {}),
                    trends_data.get('diagnosis_trends', {}),
                    trends_data.get('performance', {}),
                    header_font, border
                )

            elif report_type == 'models' and 'models' in data:
                if first_sheet:
                    ws = wb.active
                    ws.title = "Models"
                    first_sheet = False
                else:
                    ws = wb.create_sheet("Models")
                self._add_models_sheet(ws, data['models'], header_font, border)

            elif report_type == 'users' and 'users' in data:
                if first_sheet:
                    ws = wb.active
                    ws.title = "Users"
                    first_sheet = False
                else:
                    ws = wb.create_sheet("Users")
                self._add_users_sheet(ws, data['users'], header_font, border)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _add_overview_sheet(self, ws, overview_data: Dict, header_font: Font, border: Border,
                           start_date: Optional[date], end_date: Optional[date]):
        """添加概览Sheet"""
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")

        ws['A1'] = "Overview Report"
        ws['A1'].font = Font(size=16, bold=True, color="1E40AF")
        ws.merge_cells('A1:B1')

        row = 3
        if start_date or end_date:
            ws[f'A{row}'] = "Date Range:"
            ws[f'B{row}'] = f"{start_date or 'N/A'} - {end_date or 'N/A'}"
            row += 1
        ws[f'A{row}'] = "Generated:"
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        row += 2

        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Value"
        ws[f'A{row}'].fill = header_fill
        ws[f'B{row}'].fill = header_fill
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'].font = header_font
        row += 1

        metrics = [
            ("Total Cases", overview_data.get('total_cases', 0)),
            ("Total Diagnoses", overview_data.get('total_diagnoses', 0)),
            ("Avg Execution Time (s)", f"{overview_data.get('avg_execution_time_ms', 0) / 1000:.2f}"),
            ("Completion Rate (%)", f"{overview_data.get('completion_rate', 0):.1f}"),
            ("Active Users", overview_data.get('active_users', 0)),
        ]

        for metric, value in metrics:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    def _add_demographics_sheet(self, ws, demographics_data: Dict, header_font: Font, border: Border):
        """添加人口统计Sheet"""
        header_fill_blue = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        header_fill_green = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")

        ws['A1'] = "Demographics Report"
        ws['A1'].font = Font(size=16, bold=True, color="1E40AF")

        # 年龄分布
        row = 3
        ws[f'A{row}'] = "Age Distribution"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        ws[f'A{row}'] = "Age Group"
        ws[f'B{row}'] = "Count"
        ws[f'A{row}'].fill = header_fill_blue
        ws[f'B{row}'].fill = header_fill_blue
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'].font = header_font
        row += 1

        for item in demographics_data.get('age_distribution', []):
            ws[f'A{row}'] = item.get('age_group', '')
            ws[f'B{row}'] = item.get('count', 0)
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1

        # 性别分布
        row += 2
        ws[f'A{row}'] = "Gender Distribution"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        ws[f'A{row}'] = "Gender"
        ws[f'B{row}'] = "Count"
        ws[f'A{row}'].fill = header_fill_green
        ws[f'B{row}'].fill = header_fill_green
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'].font = header_font
        row += 1

        gender_map = {'male': 'Male', 'female': 'Female', 'other': 'Other', 'unknown': 'Unknown'}
        for item in demographics_data.get('gender_distribution', []):
            gender = item.get('gender', '')
            ws[f'A{row}'] = gender_map.get(gender.lower(), gender)
            ws[f'B{row}'] = item.get('count', 0)
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12

    def _add_trends_sheet(self, ws, case_trends: Dict, diagnosis_trends: Dict,
                         performance_data: Dict, header_font: Font, border: Border):
        """添加趋势Sheet"""
        header_fill_blue = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        header_fill_green = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")

        ws['A1'] = "Trends Report"
        ws['A1'].font = Font(size=16, bold=True, color="1E40AF")

        # 趋势数据
        row = 3
        ws[f'A{row}'] = "Case & Diagnosis Trends"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        ws[f'A{row}'] = "Date"
        ws[f'B{row}'] = "Cases"
        ws[f'C{row}'] = "Diagnoses"
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].fill = header_fill_blue
            ws[f'{col}{row}'].font = header_font
        row += 1

        case_series = {item['date']: item['count'] for item in case_trends.get('series', [])}
        diag_series = {item['date']: item['count'] for item in diagnosis_trends.get('series', [])}
        all_dates = sorted(set(case_series.keys()) | set(diag_series.keys()))

        for d in all_dates[-30:]:  # 最近30条
            ws[f'A{row}'] = d
            ws[f'B{row}'] = case_series.get(d, 0)
            ws[f'C{row}'] = diag_series.get(d, 0)
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].border = border
            row += 1

        # 性能统计
        row += 2
        ws[f'A{row}'] = "Performance Statistics"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Value"
        ws[f'A{row}'].fill = header_fill_green
        ws[f'B{row}'].fill = header_fill_green
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'].font = header_font
        row += 1

        exec_stats = performance_data.get('execution_time_stats', {})
        metrics = [
            ("Min Execution Time (s)", f"{exec_stats.get('min', 0) / 1000:.2f}"),
            ("Max Execution Time (s)", f"{exec_stats.get('max', 0) / 1000:.2f}"),
            ("Avg Execution Time (s)", f"{exec_stats.get('avg', 0) / 1000:.2f}"),
            ("Median Execution Time (s)", f"{exec_stats.get('median', 0) / 1000:.2f}"),
            ("Total Diagnoses", performance_data.get('total_diagnoses', 0)),
        ]

        for metric, value in metrics:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

    def _add_models_sheet(self, ws, model_data: Dict, header_font: Font, border: Border):
        """添加模型对比Sheet"""
        header_fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")

        ws['A1'] = "Model Comparison Report"
        ws['A1'].font = Font(size=16, bold=True, color="1E40AF")

        row = 3
        headers = ["Model Name", "Usage Count", "Avg Time (s)", "Usage %"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        row += 1

        models = model_data.get('models', [])
        total_usage = sum(m.get('usage_count', 0) for m in models)

        for model in models:
            usage = model.get('usage_count', 0)
            percentage = (usage / total_usage * 100) if total_usage > 0 else 0
            values = [
                model.get('model_name', ''),
                usage,
                f"{model.get('avg_execution_time_ms', 0) / 1000:.2f}",
                f"{percentage:.1f}%"
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
            row += 1

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12

    def _add_users_sheet(self, ws, user_data: Dict, header_font: Font, border: Border):
        """添加用户活动Sheet"""
        header_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")

        ws['A1'] = "User Activity Report"
        ws['A1'].font = Font(size=16, bold=True, color="1E40AF")

        row = 3
        headers = ["Rank", "User", "Username", "Cases", "Diagnoses", "Score"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        row += 1

        for i, user in enumerate(user_data.get('top_creators', []), 1):
            score = user.get('case_count', 0) + user.get('diagnosis_count', 0) * 2
            values = [
                i,
                user.get('full_name') or user.get('username', ''),
                user.get('username', ''),
                user.get('case_count', 0),
                user.get('diagnosis_count', 0),
                score
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
            row += 1

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 10
